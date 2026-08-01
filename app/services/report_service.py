"""
Service for generating reports.
"""

from sqlalchemy.orm import Session
from app.services.project_service import ProjectService
from app.services.calculator_service import CalculatorService
from app.models.project import ProjectStatus
from datetime import datetime, timedelta, date
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import io

class ReportService:
    """Generate various reports."""

    @staticmethod
    def get_dashboard_data(db: Session) -> Dict[str, Any]:
        """Get dashboard statistics."""
        # Open projects count
        open_projects = ProjectService.get_open_projects(db)
        open_count = len(open_projects)
        
        # Today's income
        today = date.today()
        today_projects = ProjectService.search_by_date_range(db, today, today)
        today_income = 0.0
        for project in today_projects:
            if project.status == ProjectStatus.COMPLETED:
                financials = CalculatorService.calculate_project_financials(db, project.id)
                today_income += financials['total_income']
        
        # Current month income
        start_of_month = date(today.year, today.month, 1)
        month_projects = ProjectService.search_by_date_range(db, start_of_month, today)
        month_income = 0.0
        month_profit = 0.0
        for project in month_projects:
            if project.status == ProjectStatus.COMPLETED:
                financials = CalculatorService.calculate_project_financials(db, project.id)
                month_income += financials['total_income']
                month_profit += financials['net_profit']
        
        # Debtors (customers with debt)
        debtors = []
        all_projects = ProjectService.get_all(db)
        for project in all_projects:
            if project.status == ProjectStatus.COMPLETED:
                financials = CalculatorService.calculate_project_financials(db, project.id)
                if financials['customer_debt'] > 0:
                    debtors.append({
                        'customer_name': project.customer.name,
                        'customer_phone': project.customer.phone,
                        'amount': financials['customer_debt'],
                        'project_id': project.id
                    })
        
        return {
            'open_projects_count': open_count,
            'today_income': today_income,
            'month_income': month_income,
            'month_profit': month_profit,
            'debtors': debtors
        }

    @staticmethod
    def generate_excel_report(db: Session, start_date: date, end_date: date) -> bytes:
        """Generate Excel report for a date range."""
        projects = ProjectService.search_by_date_range(db, start_date, end_date)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "گزارش پروژه‌ها"
        
        # Headers
        headers = ['شناسه', 'مشتری', 'تلفن', 'نوع', 'وضعیت', 'درآمد کل', 'سود خالص', 'طلب من', 'بدهی مشتری']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for project in projects:
            financials = CalculatorService.calculate_project_financials(db, project.id)
            ws.append([
                project.id,
                project.customer.name,
                project.customer.phone,
                project.project_type.value,
                project.status.value,
                financials['total_income'],
                financials['net_profit'],
                financials['my_debt'],
                financials['customer_debt']
            ])
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_pdf_report(db: Session, start_date: date, end_date: date) -> bytes:
        """Generate PDF report for a date range."""
        projects = ProjectService.search_by_date_range(db, start_date, end_date)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title = Paragraph(f"گزارش پروژه‌ها از {start_date} تا {end_date}", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Table data
        data = [['شناسه', 'مشتری', 'نوع', 'وضعیت', 'درآمد', 'سود خالص']]
        for project in projects:
            financials = CalculatorService.calculate_project_financials(db, project.id)
            data.append([
                str(project.id),
                project.customer.name,
                project.project_type.value,
                project.status.value,
                f"{financials['total_income']:,.0f}",
                f"{financials['net_profit']:,.0f}"
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return buffer.getvalue()

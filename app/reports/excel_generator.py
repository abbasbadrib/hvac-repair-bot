"""
Excel report generator.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

class ExcelReportGenerator:
    """Generate Excel reports."""
    
    @staticmethod
    def generate_report(title: str, headers: list, data: list,
                       summary: dict = None, sheet_name: str = "Report") -> bytes:
        """
        Generate an Excel report.
        
        Args:
            title: Report title
            headers: List of column headers
            data: List of rows (each row is a list)
            summary: Dictionary of summary data
            sheet_name: Name of the sheet
        
        Returns:
            bytes: Excel file as bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Title
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = 4
        for data_row in data:
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        
        # Summary
        if summary:
            row += 2
            ws.merge_cells(f'A{row}:{get_column_letter(len(headers))}{row}')
            summary_cell = ws.cell(row=row, column=1)
            summary_cell.value = "خلاصه گزارش"
            summary_cell.font = Font(bold=True, size=14)
            summary_cell.alignment = Alignment(horizontal='center')
            
            row += 1
            for key, value in summary.items():
                ws.cell(row=row, column=1).value = key
                ws.cell(row=row, column=2).value = value
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row_cell in ws[column_letter]:
                if row_cell.value:
                    max_length = max(max_length, len(str(row_cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_cells in ws.iter_rows(min_row=3, max_row=row-1, 
                                      min_col=1, max_col=len(headers)):
            for cell in row_cells:
                cell.border = thin_border
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
